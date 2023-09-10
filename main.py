from fastapi import FastAPI
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from starlette.responses import FileResponse
from starlette.exceptions import HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from cryptography.fernet import Fernet
from PyPDF2 import PdfReader, PdfWriter
from os.path import exists
from os.path import join, dirname
from dotenv import load_dotenv
from pymysqlpool.pool import Pool
import pymysql.cursors
from woocommerce import API

from PIL import Image


import csv
import pandas as pd
import os
import openpyxl 

import migrate2woo as woo

wcapi = API(
    url="https://www.shopmanual2you.com/",
    consumer_key=os.environ.get("WC_CK"),
    consumer_secret=os.environ.get("WC_CS"),
    version="wc/v3",
    timeout=60
)

dotenv_path = join(dirname(__file__), '.env')
load_dotenv(dotenv_path)

app = FastAPI()

app.mount("/previews", StaticFiles(directory="previews"), name="previews")


gauth = GoogleAuth()
gauth.CommandLineAuth()
drive = GoogleDrive(gauth)

# print(os.environ.get("APP_KEY"))

key = os.environ.get("APP_KEY")
domain = os.environ.get("DOMAIN_NAME")
db_host = os.environ.get("DB_HOST")
db_user = os.environ.get("DB_USER")
db_password = os.environ.get("DB_PASSWORD")
db_name = os.environ.get("DB_NAME")


fernet = Fernet(key)

brands = {}
categories = {}
booktypes = {}
languages = {}


def getfile(file_id):
    generate_filename = 'downloads/%s.pdf' % (file_id)
    file = drive.CreateFile({'id': file_id})
    file.GetContentFile(generate_filename)

def getimage(file_id):
    generate_filename = '/var/www/wordpress/gimages/%s.jpg' % (file_id)
    watermask_filename = 'images/watermark.png'
    file_exists = exists(generate_filename)
    if not file_exists :
        file = drive.CreateFile({'id': file_id})
        file.GetContentFile(generate_filename)
        image = Image.open(generate_filename)
        image_watermask = Image.open(watermask_filename)
        width_of_watermark , height_of_watermark = image_watermask.size
        width,height = image.size
        position = (int(width/2-width_of_watermark/2),int(height/2-height_of_watermark/2))
        image.paste(image_watermask,position,image_watermask)
        image.save(generate_filename)

def encryptcp(fileb64):
    return fernet.encrypt(fileb64.encode()).decode()

def decryptcp(fileb64):
    return fernet.decrypt(fileb64.encode()).decode()

def getIndexes():
    try:
        cursor = connection_db.cursor()
        sql = "SELECT * FROM brands"
        cursor.execute(sql)
        result_brands = cursor.fetchall()
        for brand in result_brands:
            brands[brand['name'].upper()]= {"code":brand['code'],"en":brand["ref"],"th":brand["ref_th"],"tag_en":brand["tag_ref"],"tag_th":brand["tag_ref_th"]}

        sql = "SELECT * FROM categories"
        cursor.execute(sql)
        result_categories = cursor.fetchall()
        for category in result_categories:
            categories[category['name'].upper()]= {"code":category['code'],"en":category["ref"],"th":category["ref_th"]}

        sql = "SELECT * FROM book_types"
        cursor.execute(sql)
        result_booktypes = cursor.fetchall()
        for booktype in result_booktypes:
            booktypes[booktype['name'].upper()]= {"code":booktype['code'],"en":booktype['ref'],"th":booktype['ref_th']}

        sql = "SELECT * FROM languages"
        cursor.execute(sql)
        result_languages = cursor.fetchall()
        for language in result_languages:
            languages[language['name'].upper()]= language['code']
    except:
        print("connection db problem")

try:
    pool = Pool(host=db_host, port=3306, user=db_user, password=db_password, db=db_name,autocommit=True,ping_check=True,cursorclass=pymysql.cursors.DictCursor)
    connection_db = pool.get_conn()
    getIndexes()
    woo.connection_db = connection_db
except :
    print("Error while connecting to MySQL using Connection pool ")

@app.get("/download/{fileb64}")
async def download(fileb64):
    try:
        file_id = decryptcp(fileb64).split("download")[1]
        generate_filename = 'downloads/%s.pdf' % (file_id)
        file_exists = exists(generate_filename)
        if file_exists:
            return FileResponse(generate_filename, media_type='application/octet-stream', filename=generate_filename.split("/")[-1])

        getfile(file_id)

        file_exists = exists(generate_filename)

        if not file_exists:
            raise HTTPException(status_code=404, detail="Item not found")

        return FileResponse(generate_filename, media_type='application/octet-stream', filename=generate_filename.split("/")[-1])

    except:
        raise HTTPException(status_code=404, detail="Item not found")

@app.get("/preview/{fileb64}")
async def preview(fileb64,start: int = 1, end: int = 5 ):
    try:
        file_id = decryptcp(fileb64).split("preview")[1]

        download_filename = 'downloads/%s.pdf' % (file_id)
        out_filename = 'previews/%s.pdf' % (file_id)
        watermask_filename = 'images/watermark.pdf'

        file_preview_exists = exists(out_filename)

        def iterfile():
            with open(out_filename, mode="rb") as file_like:
                yield from file_like

        if file_preview_exists:
            return StreamingResponse(iterfile(), media_type="application/pdf")

        file_download_exists = exists(download_filename)
        if not file_download_exists:
            getfile(file_id)

        pdf = PdfReader(open(download_filename, "rb"))
        pdf_writer = PdfWriter()
        watermask = PdfReader(open(watermask_filename, "rb"))
        watermask_page = watermask.pages[0]
        #pages = 5

        #if len(pdf.pages) < pages:
        #    pages = len(pdf.pages)

        for page in range((start-1),end):
            content_page = pdf.pages[page]
            content_page.merge_page(watermask_page)
            pdf_writer.add_page(content_page)
        
        pdf_writer.add_metadata(pdf.metadata)
        pdf_writer.write(out_filename)

        return StreamingResponse(iterfile(), media_type="application/pdf")
    except:
        raise HTTPException(status_code=404, detail="Item not found")


@app.get("/list_test/{path}")
async def lists(path):
    try:
        fieldnames = ["title", "id", "preview", "download"]
        csv_filename = 'documentlists.csv'
        xls_filename = 'documentlists.xlsx'

        fcsv = open(csv_filename, 'w')
        writer = csv.DictWriter(fcsv, fieldnames=fieldnames)
        writer.writeheader()

        file_list = drive.ListFile(
            {'q': "'{}' in parents and trashed=false".format(path)}).GetList()
        for file in file_list:
            if file['title'] == xls_filename:
                file.Delete()
            preview_file_id = encryptcp("preview"+file['id'])
            download_file_id = encryptcp("download"+file['id'])
            pdfjs_template = '[pdfjs-viewer url="https://{domain}/gdrive/preview/{file_id}" viewer_width=100% viewer_height=800px fullscreen=true download=true print=true]'.format(
                domain=domain,file_id=preview_file_id)
            download_template = 'https://{domain}/gdrive/download/{file_id}'.format(
                domain=domain,file_id=download_file_id)
            writer.writerow({"title": file['title'], "id": file['id'],
                            "preview": pdfjs_template, "download": download_template})
            # print('title: %s, id: %s' % (file['title'], file['id']))

        # sub folder and file
        folder_list = drive.ListFile(
            {'q': "'{}' in parents and trashed=false and mimeType='application/vnd.google-apps.folder'".format(path)}).GetList()

        for folder in folder_list:
            # print('title: %s, id: %s' % (folder['title'], folder['id']))
            file_list = drive.ListFile(
                {'q': "'{}' in parents and trashed=false".format(folder['id'])}).GetList()
            for file in file_list:
                preview_file_id = encryptcp("preview"+file['id'])
                download_file_id = encryptcp("download"+file['id'])
                pdfjs_template = '[pdfjs-viewer url= "https://{domain}/gdrive/preview/{file_id}" viewer_width=100% viewer_height=800px fullscreen=true download=true print=true]'.format(
                    domain=domain,file_id=preview_file_id)
                download_template = 'https://{domain}/gdrive/download/{file_id}'.format(
                    domain=domain,file_id=download_file_id)
                writer.writerow(
                    {"title": file['title'], "id": file['id'], "preview": pdfjs_template, "download": download_template})
                # print('\t title: %s, id: %s' % (file['title'], file['id']))

        fcsv.close()

        cvsDataframe = pd.read_csv(csv_filename)
        resultExcelFile = pd.ExcelWriter(xls_filename)
        cvsDataframe.to_excel(resultExcelFile, index=False)
        resultExcelFile.save()

        gfile = drive.CreateFile({'parents': [{'id': path}]})
        gfile.SetContentFile(xls_filename)
        gfile.Upload()

        return {"message": "Please check google drive file name %s" % (xls_filename)}
    except:
        raise HTTPException(status_code=404, detail="Path not found")

@app.get("/heartbeat/{path}")
async def heartbeat(path):
    try:
        file_list = drive.ListFile(
                {'q': "'{}' in parents and trashed=false".format(path)}).GetList()
        return {"message":"connected ok"}
    except:
        raise HTTPException(status_code=503, detail="Connect error")

@app.get("/lists/{path}")
async def lists(path):
    xls_filename = 'documentlists.xlsx'
    file_dict = dict()

    parent_folder_id = path
    parent_folder_dir = './'

    if parent_folder_dir[-1] != '/':
        parent_folder_dir = parent_folder_dir + '/'

    folder_queue = [parent_folder_id]
    dir_queue = [parent_folder_dir]
    cnt = 0

    while len(folder_queue) != 0:
        current_folder_id = folder_queue.pop(0)
        file_list = drive.ListFile(
            {'q': "'{}' in parents and trashed=false".format(current_folder_id)}).GetList()

        current_parent = dir_queue.pop(0)
        print(current_parent, current_folder_id)
        for file1 in file_list:
            if file1['title'] == xls_filename:
                file1.Delete()
            file_dict[cnt] = dict()
            file_dict[cnt]['id'] = file1['id']
            file_dict[cnt]['title'] = file1['title']
            file_dict[cnt]['dir'] = current_parent + file1['title']

            if file1['mimeType'] == 'application/vnd.google-apps.folder':
                file_dict[cnt]['type'] = 'folder'
                file_dict[cnt]['dir'] += '/'
                folder_queue.append(file1['id'])
                dir_queue.append(file_dict[cnt]['dir'])
                file_dict[cnt]['preview'] = ""
                file_dict[cnt]['download'] = ""
            else:
                preview_file_id = encryptcp("preview"+file1['id'])
                download_file_id = encryptcp("download"+file1['id'])
                pdfjs_template = '[pdfjs-viewer url= "https://{domain}/gdrive/preview/{file_id}" viewer_width=100% viewer_height=800px fullscreen=true download=true print=true]'.format(
                    domain=domain,file_id=preview_file_id)
                download_template = 'https://{domain}/gdrive/download/{file_id}'.format(
                    domain=domain,file_id=download_file_id)
                file_dict[cnt]['type'] = 'file'
                
                file_dict[cnt]['preview'] = pdfjs_template
                file_dict[cnt]['download'] = download_template

            cnt += 1
    
    for x in file_dict:
        if file_dict[x]["type"] == "file" :
            try:
                cursor = connection_db.cursor()
                sql = "INSERT INTO files (id,title,dir,preview,download) values (%s,%s,%s,%s,%s)"
                val = (file_dict[x]["id"],file_dict[x]["title"],file_dict[x]["dir"],file_dict[x]["preview"],file_dict[x]["download"])
                cursor.execute(sql,val)
                connection_db.commit()
            except:
               print(sql)

        
    #cvsDataframe = pd.DataFrame(file_dict).transpose().head(10)
    cvsDataframe = pd.DataFrame(file_dict).transpose()
    resultExcelFile = pd.ExcelWriter(xls_filename)
    cvsDataframe.to_excel(resultExcelFile, index=False)
    resultExcelFile.close()

    gfile = drive.CreateFile({'parents': [{'id': path}]})
    gfile.SetContentFile(xls_filename)
    gfile

    return {"message": "Please check google drive file name %s" % (xls_filename)}

@app.get("/v1/files/product")
async def file_product():
    file_id = '11OK7R6zqW7h7szSqI3F_gP4UVg-1mDps'
    generate_filename = '%s.xlsx' % (file_id)
    file = drive.CreateFile({'id': file_id})
    file.update
    file.GetContentFile(generate_filename)

    wb = openpyxl.load_workbook(generate_filename)
    sheet =  wb['Product']
    max_row = sheet.max_row
    start_row = 3
    for row in range(start_row, max_row):
        if sheet.cell(row=row,column=17).value != None and sheet.cell(row=row,column=22).value != None and sheet.cell(row=row,column=25).value != None and sheet.cell(row=row,column=24).value != None \
        and sheet.cell(row=row,column=1).value != None and sheet.cell(row=row,column=2).value != None and sheet.cell(row=row,column=3).value != None and sheet.cell(row=row,column=4).value != None \
        and sheet.cell(row=row,column=5).value != None and sheet.cell(row=row,column=6).value != None and sheet.cell(row=row,column=7).value != None and sheet.cell(row=row,column=8).value != None \
        and sheet.cell(row=row,column=9).value != None and sheet.cell(row=row,column=10).value != None and sheet.cell(row=row,column=11).value != None and sheet.cell(row=row,column=12).value != None \
        and sheet.cell(row=row,column=13).value != None and sheet.cell(row=row,column=14).value != None and sheet.cell(row=row,column=15).value != None :
            
            product = {}
            product["no"] = sheet.cell(row=row,column=1).value
            product["brand"] = sheet.cell(row=row,column=2).value.upper()
            product["categories_en"] = sheet.cell(row=row,column=4).value.upper()
            product["categories_th"] = sheet.cell(row=row,column=5).value.upper()
            product["booktype_en"] = sheet.cell(row=row,column=7).value.upper()
            product["booktype_th"] = sheet.cell(row=row,column=8).value.upper()
            product["parts_no"] = sheet.cell(row=row,column=10).value
            product["model"] = sheet.cell(row=row,column=11).value
            product["serial_no"] = sheet.cell(row=row,column=12).value
            product["page_no"] = str(int(sheet.cell(row=row,column=13).value))
            product["file_type"] = sheet.cell(row=row,column=14).value
            product["file_lang"] = sheet.cell(row=row,column=15).value.upper()
            product["price"] = sheet.cell(row=row,column=17).value
            product["price2"] = sheet.cell(row=row,column=22).value
            product["sku"] = 'S2Y-%s%s-%s%s-%s'%(brands[product["brand"]]["code"],categories[product["categories_en"]]["code"],booktypes[product["booktype_en"]]["code"],languages[product["file_lang"]],product["no"])

            product["file_download_id"] = (sheet.cell(row=row,column=24).value).split("/")[-2]
            images = (sheet.cell(row=row,column=25).value).split(",")
            product["file_image"] = list()
            for image in images:
                if (image.strip() != ""):
                    product["file_image"].append(image.split("/")[-2])
                    getimage(image.split("/")[-2])
            if product["price"] >= 0 and product["price2"] >= 0 :
                product_update(product)

    return {"message": "file product ok"}


@app.post("/v1/indexes")
async def add_index():
    getIndexes()
    print (brands)
    print (categories)
    print (booktypes)
    print (languages)
    return {"message": "indexes is ok"}

def checkValue(origin,word):
    if origin == word :
        return True
    return False

def product_update(product):
    print (product)
    cursor = connection_db.cursor()
    #sql = "select * from products where no=%s Limit 1"
    sql = "select * from products INNER JOIN woo_products ON products.no = woo_products.product_no and products.no=%s Limit 1"
    val = (product["no"])
    cursor.execute(sql,val)
    row =  cursor.fetchone()
    row_count = cursor.rowcount

    if row_count <= 0:
        val = (product["no"],product["brand"],product["categories_en"],product["categories_th"],product["booktype_en"],product["booktype_th"],product["parts_no"],product["model"],product["serial_no"],str(product["page_no"]),product["file_type"],product["file_lang"],product["price"],product["sku"],product["file_download_id"],','.join(product["file_image"]),product["price2"])
        print(val)
        cursor = connection_db.cursor()
        sql = "INSERT INTO products (no,brand,categories_en,categories_th,booktype_en,booktype_th,parts_no,model,serial_no,page_no,file_type,file_lang,price,sku,file_download_id,file_image,price2) value (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
        cursor.execute(sql,val)
        connection_db.commit()
        preview = encryptcp("preview"+product['file_download_id'])
        download = encryptcp("download"+product['file_download_id'])
        pdfjs_template = '[pdfjs-viewer url= "https://{domain}/gdrive/preview/{file_id}" viewer_width=100% viewer_height=800px fullscreen=true download=true print=true]'.format(
                    domain=domain,file_id=preview)
        download_template = 'https://{domain}/gdrive/download/{file_id}'.format(
                    domain=domain,file_id=download)
        product["download"] = download_template
        product["preview"] = pdfjs_template
        product["brands_ref"] = brands
        product["categories_ref"] = categories
        product["booktype_ref"] = booktypes
        woo.addProducts(product)
    else: 
        columns = []
        for k,y in product.items():
            print (k,y)
            if type(y) is list:
                y = ','.join(y)

            if checkValue(y,row[k]) == False:
                columns.append(k)

        if len(columns) > 0 :
            print ("record update")

            val = []
            query = "update products set "
            field = ""
            for column in columns:
                field += column+"=%s,"
                val.append(product[column])

            sql = query +  field[0:-1]+ " where no="+product["no"]
            print(sql)
            cursor.execute(sql,val)
            connection_db.commit()
            woo.updateProduct(columns,product,row)    
        
    