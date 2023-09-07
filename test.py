#from woocommerce import API
#wcapi = API(
#    url="https://www.shopmanual2you.com/",
#    consumer_key="ck_2b201992a3b7205b97b91ebaae9b74cffd0492b2",
#    consumer_secret="cs_43a4861af8b79e8d9f933c51fea31e1eb7b2af69",
#    version="wc/v3",
#    timeout=20
#)
#data= {
#        "translations": {
#                "en":"5809"
#            }
#
#    }
    

#result= wcapi.put("products/5811",data).json()
#for c in result:
#    print ("id:%s name:%s slug:%s"%(c['id'],c['name'],c['slug']))
#print (result)


import openpyxl 
generate_filename = 'x002.xlsx' 

wb = openpyxl.load_workbook(generate_filename)
sheet =  wb['Product']
max_row = sheet.max_row
start_row = 3
print (max_row)
for row in range(start_row, max_row):
    if sheet.cell(row=row,column=17).value != None and sheet.cell(row=row,column=22).value != None and sheet.cell(row=row,column=25).value != None and sheet.cell(row=row,column=24).value != None \
       and sheet.cell(row=row,column=1).value != None and sheet.cell(row=row,column=2).value != None and sheet.cell(row=row,column=3).value != None and sheet.cell(row=row,column=4).value != None \
       and sheet.cell(row=row,column=5).value != None and sheet.cell(row=row,column=6).value != None and sheet.cell(row=row,column=7).value != None and sheet.cell(row=row,column=8).value != None \
       and sheet.cell(row=row,column=9).value != None and sheet.cell(row=row,column=10).value != None and sheet.cell(row=row,column=11).value != None and sheet.cell(row=row,column=12).value != None \
       and sheet.cell(row=row,column=13).value != None and sheet.cell(row=row,column=14).value != None and sheet.cell(row=row,column=15).value != None :
    
        product = {}
        product["no"] = sheet.cell(row=row,column=1).value
        product["brand"] = sheet.cell(row=row,column=2).value.upper()
        product["categories_en"] = sheet.cell(row=row,column=4).value.upper()
        product["categories_th"] = sheet.cell(row=row,column=5).value.upper()
        product["booktype_en"] = sheet.cell(row=row,column=7).value.upper()
        product["booktype_th"] = sheet.cell(row=row,column=8).value.upper()
        product["parts_no"] = sheet.cell(row=row,column=10).value
        product["model"] = sheet.cell(row=row,column=11).value
        product["serial_no"] = sheet.cell(row=row,column=12).value
        product["page_no"] = str(int(sheet.cell(row=row,column=13).value))
        product["file_type"] = sheet.cell(row=row,column=14).value
        product["file_lang"] = sheet.cell(row=row,column=15).value.upper()
        product["price"] = sheet.cell(row=row,column=17).value
        product["price2"] = sheet.cell(row=row,column=22).value
        #product["sku"] = 'S2Y-%s%s-%s%s-%s'%(brands[product["brand"]]["code"],categories[product["categories_en"]]["code"],booktypes[product["booktype_en"]]["code"],languages[product["file_lang"]],product["no"])

        product["file_download_id"] = (sheet.cell(row=row,column=24).value).split("/")[-2]
        images = (sheet.cell(row=row,column=25).value).split(",")
        if product["price"] >= 0 and product["price2"] >= 0 :
            print(product["no"])
        
